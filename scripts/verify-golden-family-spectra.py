#!/usr/bin/env python3
"""Verify bundled GOLDEN profile samples against the original XLSX workbook."""

import argparse
import hashlib
import json
from pathlib import Path

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[1]
MANIFEST_PATH = REPO_ROOT / "data" / "golden-family-spectra-manifest.json"


def profile_digest(ci, profile):
    payload = {
        "ci": ci,
        "productNumber": profile["productNumber"],
        "reflectance": profile["reflectance"],
        "kOverS": profile["kOverS"],
    }
    encoded = json.dumps(payload, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest().upper()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("source_xlsx", type=Path, help="Original GOLDEN XLSX workbook")
    args = parser.parse_args()

    manifest = json.loads(MANIFEST_PATH.read_text(encoding="utf-8"))
    source = args.source_xlsx.resolve()
    workbook_hash = hashlib.sha256(source.read_bytes()).hexdigest().upper()
    expected_hash = manifest["source"]["workbookSha256"]
    if workbook_hash != expected_hash:
        raise SystemExit(f"Workbook SHA-256 mismatch: {workbook_hash} != {expected_hash}")

    sheet = load_workbook(source, data_only=True, read_only=True).active
    wavelengths = manifest["source"]["wavelengths"]
    reflectance_columns = manifest["source"]["reflectanceColumns"]
    k_over_s_columns = manifest["source"]["kOverSColumns"]
    header_row = manifest["source"]["headerRow"]

    for column, wavelength in zip(reflectance_columns, wavelengths):
        if sheet.cell(header_row, column).value != wavelength:
            raise SystemExit(f"Reflectance header mismatch at column {column}")
    for column, wavelength in zip(k_over_s_columns, wavelengths):
        if sheet.cell(header_row, column).value != wavelength:
            raise SystemExit(f"K/S header mismatch at column {column}")

    for ci, profile in manifest["profiles"].items():
        row = profile["spreadsheetRow"]
        if sheet.cell(row, 1).value != profile["productNumber"]:
            raise SystemExit(f"{ci}: product number mismatch at row {row}")
        if str(sheet.cell(row, 2).value).strip() != profile["productName"]:
            raise SystemExit(f"{ci}: product name mismatch at row {row}")

        reflectance = [round(float(sheet.cell(row, column).value) / 100, 4) for column in reflectance_columns]
        k_over_s = [float(sheet.cell(row, column).value) for column in k_over_s_columns]
        if reflectance != profile["reflectance"]:
            raise SystemExit(f"{ci}: reflectance samples do not match row {row}")
        if k_over_s != profile["kOverS"]:
            raise SystemExit(f"{ci}: K/S samples do not match row {row}")
        if profile_digest(ci, profile) != profile["profileSha256"]:
            raise SystemExit(f"{ci}: manifest profile digest mismatch")

    print(f"Verified {len(manifest['profiles'])} GOLDEN profiles against {source.name}")


if __name__ == "__main__":
    main()
